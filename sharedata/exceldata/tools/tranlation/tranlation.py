# -*- coding: utf-8 -*-

###############################################
# 一般翻译流程：
# 1、在unity里点击菜单 lua-> ReportUIString
# 2、在unity里点击菜单 lua-> GenerateTranlation
# 3、把 翻译表.xlsx 翻译好
# 4、上传excel表
#
#
# 首次运行要检查lua所有字符串，并将标记的文字自动加langlua
# 1、把主函数里的ReplaceLuaString 函数的注释去掉
# 2、运行该脚本
# 3、测试并上传修改过的lua文件
###############################################

import re
import os
import csv
import StringIO
import codecs
import sys
import copy
import xlrd

reload(sys)
sys.setdefaultencoding('utf8')

import openpyxl
from openpyxl import load_workbook
from openpyxl.workbook import Workbook  
from openpyxl.writer.excel import ExcelWriter
from openpyxl.styles import Alignment

###############################################
# 配置
###############################################
excel_path = r"../../excel"
client_lua_path = r"../../../client/LuaScript/UI"
server_lua_path = r"../../server_pub/script"
translation_lua_path = r"../../../client/LuaScript/Data/TranslationData.lua"
ui_string_path = r"ui_string.csv"
translation_excel_path = ur"../../excel/all/F-翻译表.xlsx"


lua_path_dict = [
  {"csv_path": r"lua_string_client.csv", "lua_path": r"../../../../client/LuaScript/UI/UIConst.lua", "mark":"客户端lua"},
  {"csv_path": r"lua_string_server.csv", "lua_path": r"../../../../server/lualib/tips.lua", "mark":"服务端lua"},
]


STRING_SPLIT = '//'
LUA_SHEET_NAME = u"Lua翻译表"
UI_SHEET_NAME = u"UI翻译表"
EXCEL_SHEET_NAME = u"Excel翻译表"

LUA_INIT_ID = 100001
UI_INIT_ID = 200001
EXCEL_INIT_ID = 300001

tokens = [
  ("comment_1" , re.compile(r"--\[\[.*?\]\]", re.S)),
  ("comment_2" , re.compile(r"--\[\=\[.*?\]\=\]", re.S)),
  ("comment_3" , re.compile(r"--\[\=\=\[.*?\]\=\=\]", re.S)),
  ("comment_0" , re.compile(r"--[^\r\n]*", re.S)),
  ("str_1" , re.compile(r'("((\\.|[^"\\])*)")', re.S)),
  ("str_2" , re.compile(r"('((\\.|[^'\\])*)')", re.S)),
  ("str_3" , re.compile(r'(\[\[(.*?)\]\])', re.S)),
  ("str_4" , re.compile(r'(\[\=\[(.*?)\]\=\])', re.S)),
  ("str_5" , re.compile(r'(\[\=\=\[(.*?)\]\=\=\])', re.S)),
]
chinese_re = re.compile(ur'[\u4E00-\u9FA5]')

###############################################
# 基础文件读写支持
###############################################
def GetLuaFileList():
  path_list = []
  for path_tb in lua_path_dict:
    if os.path.exists(path_tb["lua_path"]):
      path_list.append(path_tb)
  return path_list

def ReadFile(full_path):
  file = open(full_path, 'rb')
  content = file.read()
  file.close()
  return content

def WriteFile(full_path,content):
  file = open(full_path, 'wb')
  file.write(content)
  file.close()

def ReadCSV(filename):
  if not os.path.isfile(filename):
    return False
  lines = []
  file = open(filename, 'rb')
  reader = csv.reader(file)
  for line in reader:
    if line and line[0] and line[0] != "":
      lines.append(line)
  file.close()
  return lines

def WriteCSV(filename, lines):
  file = open(filename, 'wb')
  #file = codecs.open(filename, 'wb', "gbk") #用codecs处理中文乱码
  writer = csv.writer(file)
  for line in lines:
    writer.writerow(line)
  file.close()

def ReadExcel(filename, sheet_name, begin_row = 1):
  if not os.path.isfile(filename):
    return False
  book = load_workbook(filename, data_only=True)
  if not book.get_sheet_by_name(sheet_name):
    return False
  sheet = book[sheet_name]
  max_row = sheet.max_row
  max_column = sheet.max_column
  lines = []
  for row in xrange(begin_row, max_row+1):  #行
    line = []
    is_add = False
    for column in xrange(1, max_column+1):  #列
      value = sheet.cell(row=row, column=column).value
      if type(value) == unicode:
        value = value.encode("utf8") 
      line.append(value)
      if value:
        is_add = True
    if is_add :
      lines.append(line)
  return lines


def WriteExcel(filename, sheet_name, lines, begin_row = 1):
  if not os.path.isfile(filename):
    return False
  book = load_workbook(filename, data_only=True)
  if not book.get_sheet_by_name(sheet_name):
    return False
  sheet = book[sheet_name]
  ClearSheet(sheet, begin_row)
  style = sheet.cell(1, 1).style
  row = begin_row
  for line in lines:
    column = 1
    for field in line:
      cell = sheet.cell(row=row, column=column)
      cell.style = style
      cell.alignment = Alignment(wrapText=True)
      cell.value = field
      column+=1
    row+=1
  book.save(filename)

def ClearSheet(sheet, begin_row = 1):
  max_row = sheet.max_row
  max_column = sheet.max_column
  for row in xrange(begin_row, max_row+1): 
    for column in xrange(1, max_column+1):
      sheet.cell(row=row, column=column).value = None


###############################################
# 翻译扫描相关功能
###############################################
def ScanChineseStringInLua(content):
  matchs = {}
  for (token, r) in tokens:
    match = r.search(content)
    matchs[token] = match
  i = 0
  str_len = len(content)
  result = []
  while True:
    min_token = None
    min_i = str_len
    for (token, r) in tokens:
      match = matchs[token]
      if not match:
        continue
      if match.start() < i:
        match = r.search(content, i)
        matchs[token] = match
      if match and match.start() < min_i:
        min_token = token
        min_i = match.start()
    if min_token:
      is_translate = False
      if min_token.startswith("str"):  #匹配到字符串
        s = matchs[min_token].group(1)
        pure_str = matchs[min_token].group(2)
        start = matchs[min_token].start()
        if chinese_re.search(s.decode('utf8')):  #只翻译包含中文的
          result.append([min_token, s, pure_str, start])
    else:
      break
    i = matchs[min_token].end()
  return result

def HasChinese(str):
  return chinese_re.search(str.decode('utf8'))

def Escape(str):
    str = str.replace("\r\n", "\n") \
               .replace("\r", "\n")
    if str.startswith("\n"):
        str = str[1:]
    # str = str.replace("\\", r"\\") \
    #          .replace("\a", r"\a") \
    #          .replace("\b", r"\b") \
    #          .replace("\f", r"\f") \
    #          .replace("\n", r"\n") \
    #          .replace("\r", r"\r") \
    #          .replace("\t", r"\t") \
    #          .replace("\v", r"\v") \
    #          .replace("\"", r"\"") \
    #          .replace("\'", r"\'")
    return str

###############################################
# 功能入口
###############################################
def CollectAllTranslation():
  print "CollectAllTranslation"
  CollectLua()
  #CollectUI()
  CollectExcel()

def CollectLua():
  print "CollectLua"
  # lua翻译表 line:{mark,id,count,raw,chs,cht,eng}
  MARK_INDEX = 0
  ID_INDEX = 1
  COUNT_INDEX = 2
  RAW_INDEX = 3

  TEXT_INDEX = 0

  old_list = ReadExcel(translation_excel_path, LUA_SHEET_NAME, 7)
  if not old_list:
    old_list = []
  old_dict = {}
  new_list = []
  new_dict = {}
  line_id = LUA_INIT_ID
  print "old:", len(old_list)
  for line in old_list:
    old_dict[line[RAW_INDEX]] = line
  for path_tb in lua_path_dict:
    new_lines = ReadCSV(path_tb["csv_path"])
    if new_lines:
      for line in new_lines:
        text = Escape(line[TEXT_INDEX])
        if new_dict.has_key(text):
          new_line = new_dict[text]
          new_line[COUNT_INDEX] = new_line[COUNT_INDEX] + 1
        else:
          new_line = []
          if old_dict.has_key(text) :
            new_line = copy.copy(old_dict[text])
            new_line[MARK_INDEX] = path_tb["mark"]
            new_line[ID_INDEX] = line_id
            new_line[COUNT_INDEX] = 1
          else:
            new_line = [path_tb["mark"], line_id, 1, text]
          new_list.append(new_line)
          new_dict[text] = new_line
          line_id = line_id + 1

  for line in old_list:
      if not new_dict.has_key(line[RAW_INDEX]):
        new_line = copy.copy(line)
        new_line[ID_INDEX] = line_id
        new_line[COUNT_INDEX] = 0
        new_list.append(new_line)
        line_id = line_id + 1

  print "new:", len(new_list)
  WriteExcel(translation_excel_path, LUA_SHEET_NAME, new_list, 7)

def CollectUI():
  print "CollectUI"
  # ui翻译表 line:{mark,id,count,raw,chs,cht,eng}
  MARK_INDEX = 0
  ID_INDEX = 1
  COUNT_INDEX = 2
  RAW_INDEX = 3

  UI_NAME_INDEX = 0
  COMP_INDEX = 1
  TEXT_INDEX = 2

  old_list = ReadExcel(translation_excel_path, UI_SHEET_NAME, 7)
  if not old_list:
    old_list = []
  old_dict = {}
  new_list = []
  new_dict = {}
  line_id = UI_INIT_ID
  print "old:", len(old_list)
  for line in old_list:
    old_dict[line[RAW_INDEX]] = line
  new_lines = ReadCSV(ui_string_path)
  if new_lines:
    for line in new_lines:
      ui_name = line[UI_NAME_INDEX]
      #comp_id = line[COMP_INDEX]
      text = Escape(line[TEXT_INDEX])
      if HasChinese(text):
        if new_dict.has_key(text):
          new_line = new_dict[text]
          new_line[COUNT_INDEX] = new_line[COUNT_INDEX] + 1
          mark_text = new_line[MARK_INDEX]
          is_add = True
          for name in mark_text.split(","):
            if name == ui_name:
              is_add = False
              break
          if is_add:
            new_line[MARK_INDEX] = new_line[MARK_INDEX] + "," + ui_name
        else:
          new_line = []
          if old_dict.has_key(text):
            new_line = copy.copy(old_dict[text])
            new_line[ID_INDEX] = line_id
            new_line[COUNT_INDEX] = 1
          else:
            new_line = [ui_name, line_id, 1, text]
            #print ui_name, line_id, text
          new_list.append(new_line)
          new_dict[text] = new_line
          line_id = line_id + 1
  for line in old_list:
      if not new_dict.has_key(line[RAW_INDEX]):
        new_line = copy.copy(line)
        new_line[ID_INDEX] = line_id
        new_line[COUNT_INDEX] = 0
        new_list.append(new_line)
        line_id = line_id + 1
  print "new:", len(new_list)
  WriteExcel(translation_excel_path, UI_SHEET_NAME, new_list, 7)

def AppendLine(sheet, new_lines, column, title, is_list = False):
  field_name = sheet.cell(3, column).value
  rows = sheet.col_values(column, 6)
  for row in rows:
    if type(row) == float:
        if int(row) == row:
          row = str(int(row))
        else:
          row = str(row)
    if row.strip():
      str_list = []
      if is_list :
        str_list = row.split(STRING_SPLIT)
      else:
        str_list.append(row)
      for s in str_list:
        line = []
        line.append(title)
        line.append(field_name)
        if type(s) == unicode:
          s = s.encode("utf8")
        line.append(s)
        new_lines.append(line)

def CollectExcel():
  print "CollectExcel"
  all_xls_files = []
  new_lines = []
  for dirpath, dirnames, filenames in os.walk(excel_path):
    for filename in filenames:
      if filename.startswith("~"):
        continue
      if filename.endswith("xls") or filename.endswith("xlsx"):
        all_xls_files.append(os.path.join(dirpath, filename))
  for filename in all_xls_files:
    for sheet in xlrd.open_workbook(filename).sheets():
      if sheet.nrows == 0 or sheet.name.startswith("#"):
        continue
      title = sheet.cell(3, 0).value
      for column in xrange(0, sheet.ncols):
        val_type = sheet.cell(4, column).value
        if val_type == "lang" or val_type == "lang_list":
          AppendLine(sheet, new_lines, column, title, val_type == "lang_list")

  # excel翻译表 line:{mark,id,count,raw,chs,cht,eng}
  MARK_INDEX = 0
  ID_INDEX = 1
  COUNT_INDEX = 2
  RAW_INDEX = 3

  EXCEL_TITLE_INDEX = 0
  EXCEL_FIELD_INDEX = 1
  EXCEL_TEXT_INDEX = 2

  old_list = ReadExcel(translation_excel_path, EXCEL_SHEET_NAME, 7)
  if not old_list:
    old_list = []
  old_dict = {}
  new_list = []
  new_dict = {}
  line_id = EXCEL_INIT_ID
  print "old:", len(old_list)
  for line in old_list:
    old_dict[line[RAW_INDEX]] = line
  if new_lines:
    for line in new_lines:
      title = line[EXCEL_TITLE_INDEX]
      text = Escape(line[EXCEL_TEXT_INDEX])
      if new_dict.has_key(text):
        new_line = new_dict[text]
        new_line[COUNT_INDEX] = new_line[COUNT_INDEX] + 1
        mark_text = new_line[MARK_INDEX]
        is_add = True
        for name in mark_text.split(","):
          if name == title:
            is_add = False
            break
        if is_add:
          new_line[MARK_INDEX] = new_line[MARK_INDEX] + "," + title
      else:
        new_line = []
        if old_dict.has_key(text):
          new_line = copy.copy(old_dict[text])
          new_line[ID_INDEX] = line_id
          new_line[COUNT_INDEX] = 1
        else:
          new_line = [title, line_id, 1, text]
        new_list.append(new_line)
        new_dict[text] = new_line
        line_id = line_id + 1
  for line in old_list:
      if not new_dict.has_key(line[RAW_INDEX]):
        new_line = copy.copy(line)
        new_line[ID_INDEX] = line_id
        new_line[COUNT_INDEX] = 0
        new_list.append(new_line)
        line_id = line_id + 1
  print "new:", len(new_list)
  WriteExcel(translation_excel_path, EXCEL_SHEET_NAME, new_list, 7)



def CollectLuaTranslation():
  print "CollectLuaTranslation"
  lines = []
  file_list = GetLuaFileList()
  for path_tb in file_list:
    content = ReadFile(path_tb["lua_path"])
    results = ScanChineseStringInLua(content)
    for result in results:
      s = result[2]
      start = result[3]
      if content[start - 8:start] == "langlua[":
          lines.append([s])
    print path_tb["lua_path"], "length:", len(lines) 
    WriteCSV(path_tb["csv_path"], lines)

def ReplaceLuaString():
  print "ReplaceLuaString"
  file_list = GetLuaFileList()
  for path_tb in file_list:
    content = ReadFile(path_tb["lua_path"])
    results = ScanChineseStringInLua(content)
    new_content = []
    last_end = 0
    for result in results:
      s = result[1]
      start = result[3]
      if content[start - 8:start] != "langlua[":
        new_content.append(content[last_end:start])
        new_content.append("langlua[" + s + "]") 
        last_end = start + len(s)
    new_content.append(content[last_end:])
    WriteFile(path_tb["lua_path"], "".join(new_content))


###############################################
# 主函数
###############################################
if __name__ == '__main__':
  #ReplaceLuaString()  # 把需要翻译的字符串加上langlua[]
  CollectLuaTranslation()  # 收集lua里所有的langlua[]
  CollectAllTranslation()  # 收集所有需要翻译的字符串
